import sqlite3
import openpyxl
import  time

lists=sqlite3.connect('bus.db')
c=lists.cursor()

# SQL = '''
# CREATE TABLE %s (
# 序号 INT PRIMARY KEY NOT NULL,
# 车次 TEXT NOT NULL,
# 海关 TEXT,
# 环卫公司 TEXT,
# 普瑞华庭 TEXT,
# 洋浦大桥北 TEXT,
# 信立花园 TEXT,
# 铁炉村 TEXT,
# 安泰小区 TEXT,
# 公安局 TEXT,
# 法院 TEXT,
# 浦馨苑北 TEXT,
# 阳光海湾 TEXT,
# 检察院 TEXT,
# 洋浦检疫局 TEXT,
# 蓝领公寓二期 TEXT,
# 洋浦技工学校 TEXT,
# 公交公司 TEXT,
# 新都社区 TEXT,
# 新都工业园 TEXT,
# 高山社区 TEXT,
# 金海北门 TEXT,
# 金海生活区 TEXT,
# 公堂下村 TEXT,
# 海南LNG TEXT,
# 中石化香港 TEXT,
# 海南逸盛 TEXT,
# 中石化油储 TEXT,
# 华信孚宝 TEXT,
# 山东高速 TEXT
# )
# ''' % ("bus_table")
# c.execute(SQL)
#
# listinsheet = openpyxl.load_workbook(r'3bus.xlsx')
# datainlist = listinsheet.active
# data_truck = '''INSERT INTO bus_table(序号,车次,海关,环卫公司,普瑞华庭,洋浦大桥北,信立花园,铁炉村,安泰小区,公安局,法院,浦馨苑北,阳光海湾,检察院,洋浦检疫局,蓝领公寓二期,洋浦技工学校,公交公司,新都社区,新都工业园,高山社区,金海北门,金海生活区,公堂下村,海南LNG,中石化香港,海南逸盛,中石化油储,华信孚宝,山东高速) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)'''
#
# for row in datainlist.iter_rows(min_row = 2, max_col = 30, max_row = datainlist.max_row):
#     cargo = [str(cell.value) for cell in row]
#     c.execute(data_truck, cargo)
# lists.commit()


c.execute("select 中石化香港 from bus_table")

x = c.fetchall()
x = x[2:]

for i in x:
    print(i[0])
    
lists.close()
